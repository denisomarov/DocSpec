# подключаем библиотеки

import os, shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.styles.numbers import BUILTIN_FORMATS

# устанавливаем переменные для работы
PATH = './output files/KMG'                                           # путь к папке или файлу с данными
SAVE_PATH = './output files/'                                         # путь к папке для сохранения результатов
SAVE_FILE = './output files/Расчет цены ПНР КМГ.xlsx'                 # путь к файлу для сохранения результата
TEMPLATE_FILE = './output files/Base template. Ver 2.xlsx'            # путь к файлу шаблона для сохранения результата

# формируем перечень файлов для извлечения данных
file_list = []
folder = PATH

# формируем список файлов

for root, dirs, files in os.walk(folder):
    for file in files:
        if file.endswith('xlsx') and not file.startswith('~'):
            file_list.append(os.path.join(root, file))

# извлекаем данные из файлов и формируем файл с уникальными значениями

for file in file_list:
    data_file = pd.read_excel(file, sheet_name=0, header=0)
    # data_file.columns = ['№ п/п.', 'Наименование', 'Тип', 'Ед. изм', 'Кол-во']

    if len(data_file) > 0:
        # очищаем данные от пустых ячеек
        data_file.dropna()
        if file == file_list[0]:
            data = data_file[['Наименование', 'Тип', 'Ед. изм', 'Кол-во']]
        else:
            data = pd.concat([data[['Наименование', 'Тип', 'Ед. изм', 'Кол-во']], \
                              data_file[['Наименование', 'Тип', 'Ед. изм', 'Кол-во']]], ignore_index=True)

    else:
        print('Пустой файл ', file)

# записываем итоговую таблицу в файл
# подсчитываем суммарное значение приборов по типу

data = data.groupby(by=["Тип"]).sum().reset_index()

# готовим файл для заполнения - копируем и переименовываем шаблон

shutil.copy(TEMPLATE_FILE, SAVE_FILE)

# data.to_excel(SAVE_PATH+'Итог.xlsx', header=['Тип', 'Кол-во'], \
#               sheet_name='Общая спецификация', index = False, engine='openpyxl')

# заполняем столбец Кол-во файла для сохранения результата
# открываем файл для изменения стиля ячеек

wb = load_workbook(SAVE_FILE)
sheet = wb['Спецификация']

# проводим перебор ячеек

for i in range(1, sheet.max_row):

    if data.loc[data['Тип']==sheet['C' + str(i)].value, 'Кол-во'].values:
        sheet['E' + str(i)].value = data.loc[data['Тип']==sheet['C' + str(i)].value, 'Кол-во'].values[0]

# записываем результат в файл

wb.save(SAVE_FILE)