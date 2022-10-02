# Формируем таблицу уникальных значений типов оборудования
# которые нужно преобразовать в принятые типы, то есть
# в Word-файле тип оборудования может быть записан
# отлично от единого, принятого стандартизированного образца
# и мы это обнаруживаем.
# Далее, файл будет дополнен в ручную стандартизированным
# обозначение и будет использован для замены значений
# стандартизированными значениями.

# подключаем библиотеки
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.styles.numbers import BUILTIN_FORMATS

# устанавливаем переменные для работы
PATH = './output files/JMG'                                           # путь к папке или файлу с данными
SAVE_PATH = './output files/'                                         # путь к папке для сохранения результата

# формируем перечень файлов для извлечения данных
file_list = []
folder = PATH

# формируем список файлов

for root, dirs, files in os.walk(folder):
    for file in files:
        if file.endswith('xlsx') and not file.startswith('~'):
            file_list.append(os.path.join(root, file))

# извлекаем данные из файлов и формируем файл с уникальными значениями

for file in file_list:
    data_file = pd.read_excel(file, sheet_name=0, header=0)
    if len(data_file) > 0:
        # очищаем данные от пустых ячеек
        data_file.dropna()
        if file == file_list[0]:
            data = data_file[['Наименование', 'Тип', 'Ед. изм']]
        else:
            data = pd.concat([data[['Наименование', 'Тип', 'Ед. изм']], \
                              data_file[['Наименование', 'Тип', 'Ед. изм']]], ignore_index=True)

    else:
        print('Пустой файл ', file)

# записываем итоговую таблицу в файл

data.to_excel(SAVE_PATH+'Base price.xlsx', header=['Наименование', 'Тип', 'Ед. изм'], \
              sheet_name='Базовая цена', index = False, engine='openpyxl')

# # преобразуем файл Excel
# # открываем файл для изменения стиля ячеек
#
# wb = load_workbook(SAVE_PATH+'Base price.xlsx')
# sheet = wb['Базовая цена']
#
# # фиксируем стиль для границ ячейки (рамку)
#
# bd = Side(style='thin', color="000000")
#
# # проводим перебор ячеек и меняем стили
#
# col_name = 'ABC'
#
# for k in range(0, 4):
#     for i in range(1, len(data) + 2):
#         sheet[col_name[k] + str(i)].font = Font(name='Calibri', bold=False, size=11)
#         sheet[col_name[k] + str(i)].border = Border(left=bd, top=bd, right=bd, bottom=bd)
#         # sheet[col_name[k] + str(i)].alignment = Alignment(wrap_text=True)
#
#         # изменение формата ячейки для строковых значений
#
#         if k == 0:
#             sheet[col_name[k] + str(i)].alignment = Alignment(horizontal='left', vertical='center', \
#                                                               wrap_text=True)
#
#         # изменение формата ячейки для строковых значений
#
#         if (k == 1) or (k == 2):
#             sheet[col_name[k] + str(i)].alignment = Alignment(horizontal='center', vertical='center', \
#                                                               wrap_text=True)
#         # # изменение формата ячейки для целых чисел
#         #
#         # if k == 3:
#         #     sheet[col_name[k] + str(i)].number_format = BUILTIN_FORMATS[1]
#         #     sheet[col_name[k] + str(i)].alignment = Alignment(horizontal='center', vertical='center', \
#         #                                                       wrap_text=True)
#
#         # изменение формата ячейки для денежных значений
#
#         # if (k == 5) or (k == 6):
#         #     sheet[col_name[k] + str(i)].number_format = BUILTIN_FORMATS[4]
#         #     sheet[col_name[k] + str(i)].alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
#
#         # изменение формата ячеек первой строки
#
#         if (i == 1):
#             sheet[col_name[k] + str(i)].fill = PatternFill("solid", fgColor="00AAEE")
#             sheet[col_name[k] + str(i)].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
#             sheet[col_name[k] + str(i)].font = Font(bold=True)
#
# # изменяем ширину столбцов
#
# sheet.column_dimensions['A'].width = 40
# sheet.column_dimensions['B'].width = 25
# sheet.column_dimensions['C'].width = 10
#
# # записываем результат в файл
#
# wb.save(SAVE_PATH+'Base price.xlsx')
