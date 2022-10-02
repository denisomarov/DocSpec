# Работаем с алиасами подмены - ищем и заменяем названия приборов
# в имеющихся спецификациях, так как спецификации собираются из
# Word-файлов и других источников, то может иметь место неверное
# написание приборов. Этот модуль вносит исправления и стандартизирует
# спецификации на каждый объект мр Каламкасмунайгаз

# подключаем библиотеки
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.styles.numbers import BUILTIN_FORMATS

# устанавливаем переменные для работы
PATH = './output files/KMG/'                                        # путь к папке или файлу с данными
SAVE_PATH = './output files/'                                        # путь к папке для сохранения результата
SAVE_FILE = './output files/Расчет цены ПНР КМГ.xlsx'                # путь к папке для сохранения результата
ALIAS_FILE = './output files/Список замены КМГ. Ред 2.xlsx'          # путь к папке для сохранения результата

# формируем таблицу с алиасами подмены

alias_names = pd.read_excel(ALIAS_FILE, sheet_name='alias', header=0, engine='openpyxl')

# формируем перечень файлов для извлечения данных
file_list = []
folder = PATH

# формируем список файлов

for root, dirs, files in os.walk(folder):
    for file in files:
        if file.endswith('xlsx') and not file.startswith('~'):
            file_list.append(os.path.join(root, file))

# извлекаем данные из файлов и формируем файл с уникальными значениями

for file in file_list:
    data_file = pd.read_excel(file, sheet_name=0, header=0)
    data_file.columns = ['№ п/п.', 'Наименование', 'Тип', 'Ед. изм', 'Кол-во']

    if len(data_file) > 0:
        # соединяем данные из файла и алиасы
        data = data_file.merge(alias_names, how='left', on='Тип')

    # обработка полученных данных и запись в файл

    data['Наименование'] = data['Наименование базовое']
    data['Тип'] = data['Марка, тип']
    data = data[['Наименование', 'Тип', 'Ед. изм', 'Кол-во']]
    data = data.dropna()

    data.to_excel(file, sheet_name='Спецификация', index = False, engine='openpyxl')

    # преобразуем файл Excel
    # открываем файл для изменения стиля ячеек

    wb = load_workbook(file)
    sheet = wb['Спецификация']

    # фиксируем стиль для границ ячейки (рамку)

    bd = Side(style='thin', color="000000")

    # проводим перебор ячеек и меняем стили

    col_name = 'ABCD'

    for k in range(0, 4):
        for i in range(1, len(data) + 2):
            sheet[col_name[k] + str(i)].font = Font(name='Calibri', bold=False, size=11)
            sheet[col_name[k] + str(i)].border = Border(left=bd, top=bd, right=bd, bottom=bd)
            # sheet[col_name[k] + str(i)].alignment = Alignment(wrap_text=True)

            # изменение формата ячейки для строковых значений

            if k == 0:
                sheet[col_name[k] + str(i)].alignment = Alignment(horizontal='left', vertical='center', \
                                                                  wrap_text=True)

            # изменение формата ячейки для строковых значений

            if (k == 1) or (k == 2):
                sheet[col_name[k] + str(i)].alignment = Alignment(horizontal='center', vertical='center', \
                                                                  wrap_text=True)
            # изменение формата ячейки для целых чисел

            if k == 3:
                sheet[col_name[k] + str(i)].number_format = BUILTIN_FORMATS[1]
                sheet[col_name[k] + str(i)].alignment = Alignment(horizontal='center', vertical='center', \
                                                                  wrap_text=True)

            # изменение формата ячеек первой строки

            if (i == 1):
                sheet[col_name[k] + str(i)].fill = PatternFill("solid", fgColor="00AAEE")
                sheet[col_name[k] + str(i)].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet[col_name[k] + str(i)].font = Font(bold=True)

    # изменяем ширину столбцов

    sheet.column_dimensions['A'].width = 40
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 10

    # записываем результат в файл

    wb.save(file)